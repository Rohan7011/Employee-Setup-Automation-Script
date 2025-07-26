import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Emp Setup Request Sheet Extractor", layout="centered")
st.title("Employee Setup Task Extractor")

st.markdown("""
Upload the Excel file received from Crystal Reports.
The app will extract and filter rows where **Task Desc = 'Emp Setup 08.1- SAP Primary Role'** and return relevant columns only.
""")

uploaded_file = st.file_uploader("Upload Excel file", type=["xls", "xlsx"])

if uploaded_file is not None:
    try:
        # Read Excel based on extension
        if uploaded_file.name.endswith('.xls'):
            raw0 = pd.read_excel(uploaded_file, header=None, dtype=str, engine='xlrd')
        else:
            raw0 = pd.read_excel(uploaded_file, header=None, dtype=str)
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
    else:
        # 1. Locate main and subheader rows
        row_main = row_sub = None
        for i in range(len(raw0)):
            vals = raw0.iloc[i].fillna('').astype(str).str.strip().tolist()
            if row_main is None and 'HD ID' in vals:
                row_main = i
            if row_sub is None and 'Task ID' in vals:
                row_sub = i
            if row_main is not None and row_sub is not None:
                break

        if row_main is None or row_sub is None:
            st.error("Could not find both 'HD ID' and 'Task ID' header rows.")
            st.dataframe(raw0.head(10))
        else:
            # 2. Build headers
            main_hdr = raw0.iloc[row_main].fillna('').astype(str).str.strip().tolist()
            sub_hdr = raw0.iloc[row_sub].fillna('').astype(str).str.strip().tolist()
            flat_cols = [sub_hdr[c] if sub_hdr[c] else main_hdr[c] for c in range(raw0.shape[1])]

            # 3. Slice and set columns
            data = raw0.iloc[row_sub+1:].copy()
            data.columns = flat_cols

            # 4. Clean and filter
            data['HD ID'] = data['HD ID'].replace('HD ID', np.nan)
            data['HD ID'] = data['HD ID'].ffill()

            expected = ['HD ID', 'Task ID', 'Task Desc', 'Task Tech', 'Task Create', 'Task Status', 'Task Group']
            missing = [col for col in expected if col not in data.columns]

            if missing:
                st.error(f"Missing columns in file: {missing}")
                st.write("Detected columns:", data.columns.tolist())
            else:
                clean = data.dropna(how='all')
                clean = clean[clean['Task Desc'] == "Emp Setup 08.1- SAP Primary Role"]
                clean = clean[expected]

                if clean.empty:
                    st.warning("No matching 'Emp Setup 08.1- SAP Primary Role' rows found.")
                else:
                    # Create downloadable Excel file
                    output = BytesIO()
                    clean.to_excel(output, index=False)
                    output.seek(0)

                    # Add filters
                    wb = load_workbook(output)
                    ws = wb.active
                    last_col = ws.max_column
                    last_col_letter = get_column_letter(last_col)
                    ws.auto_filter.ref = f"A1:{last_col_letter}1"

                    # Save again
                    final_output = BytesIO()
                    wb.save(final_output)
                    final_output.seek(0)

                    st.success("Filtered file ready. Download below:")
                    st.download_button(
                        label="Download Extracted Excel",
                        data=final_output,
                        file_name=f"EmpSetup_Requests_{datetime.today():%Y-%m-%d}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
